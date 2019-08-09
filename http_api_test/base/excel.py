from openpyxl import load_workbook
from base import logger

_HEADER = ['id', 'name', 'ignore', 'method', 'url', 'params', 'headers', 'body', 'timeout', 'statuscode', 'checkpoints',
           'setprops']


def read_sheet(workbook, sheet):
    """
    从workbook中读取一个sheet
    :param workbook: workbook的对象
    :param sheet: sheet名字的str
    :return: [{记录用例信息的字典},...]
    """
    logger.debug('sheet: '+sheet)
    tcs = []
    # 判断一下i.value是不是None，为了防止将空的单元格也读进来
    first_line = [i.value for i in next(workbook[sheet].iter_rows(min_row=1, max_row=1)) if i.value is not None]
    if first_line != _HEADER:
        logger.error('测试用例表格的格式错误，不读取该sheet！')
        return []

    for row in workbook[sheet].iter_rows(min_row=2):
        tc = {}
        for i, cell in zip(range(len(_HEADER)), row):
            tc[_HEADER[i]] = cell.value
        tcs.append(tc)
    return tcs


def read_wb(workbook, sheet=None):
    """
    :param workbook: workbook路径的str
    :param sheet: sheet名字的str，如果是None，则读取整个workbook
    :return: [{记录用例信息的字典},...]
    """
    logger.debug('workbook: '+workbook)
    wb = load_workbook(workbook)
    tcs = []
    if sheet is not None:
        tcs.extend(read_sheet(wb, sheet))
    else:
        for i in wb.sheetnames:
            tcs.extend(read_sheet(wb, i))
    return tcs


def read_wbs(tcs_path):
    """
    :param tcs_path: 元素可以是二元组或一元组的列表
    :return: [{记录用例信息的字典},...]
    """
    logger.debug('tcs_path: '+str(tcs_path))
    tcs = []
    for i in tcs_path:
        if len(i) != 1 and len(i) != 2:
            logger.warning('序列的长度错误: ' + str(i))
            continue
        tcs.extend(read_wb(i[0], None if len(i) == 1 else i[1]))

    return tcs
